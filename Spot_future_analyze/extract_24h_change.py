#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import time
from typing import Any, Dict, List, Optional

import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule


# =========================
# 配置
# =========================
INPUT_FILE = r"./data/24h_change_table.xlsx"
OUTPUT_FILE = r"./data/binance_symbol_analysis_中文版.xlsx"
INPUT_SHEET_NAME = "24h_change_table"

SPOT_BASE = "https://api.binance.com"
FUTURES_BASE = "https://fapi.binance.com"
TIMEOUT = 20

session = requests.Session()
session.headers.update({
    "User-Agent": "binance-symbol-analysis/1.3"
})


# =========================
# 工具函数
# =========================
def safe_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        x = x.strip()
        if not x:
            return None
        try:
            return float(x)
        except ValueError:
            return None
    return None


def pct_change(now_value: Optional[float], old_value: Optional[float]) -> Optional[float]:
    if now_value is None or old_value is None or old_value == 0:
        return None
    return (now_value - old_value) / old_value


def format_large_number(value: Any) -> Optional[str]:
    """
    将大额数字简写为:
    1250 -> 1.25k
    10420000 -> 10.42m
    3450000000 -> 3.45b
    """
    num = safe_float(value)
    if num is None:
        return None

    abs_num = abs(num)

    if abs_num >= 1_000_000_000:
        return f"{num / 1_000_000_000:.2f}b"
    elif abs_num >= 1_000_000:
        return f"{num / 1_000_000:.2f}m"
    elif abs_num >= 1_000:
        return f"{num / 1_000:.2f}k"
    else:
        return f"{num:.2f}"


def format_price(value: Any) -> Optional[str]:
    num = safe_float(value)
    if num is None:
        return None

    if num >= 1000:
        return f"{num:.2f}"
    elif num >= 1:
        return f"{num:.4f}"
    else:
        return f"{num:.8f}"


def get_json(url: str, params: Optional[Dict[str, Any]] = None, max_retries: int = 5) -> Any:
    for attempt in range(max_retries):
        resp = session.get(url, params=params, timeout=TIMEOUT)

        if resp.status_code in (429, 418):
            retry_after = resp.headers.get("Retry-After")
            wait_sec = float(retry_after) if retry_after else min(2 ** attempt, 30)
            print(f"[rate-limit] {url} -> sleep {wait_sec}s")
            time.sleep(wait_sec)
            continue

        resp.raise_for_status()
        return resp.json()

    raise RuntimeError(f"Failed after retries: {url}")


def normalize_symbol(symbol_text: str) -> str:
    s = str(symbol_text).strip().upper()
    if not s:
        return s
    if s.endswith("USDT"):
        return s
    return f"{s}USDT"


# =========================
# 读取 Excel
# =========================
def read_symbols_from_excel(file_path: str, sheet_name: str) -> List[Dict[str, str]]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")

    ws = wb[sheet_name]

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if "symbol" not in headers:
        raise ValueError("未找到 'symbol' 列")

    symbol_col_idx = headers.index("symbol") + 1

    rows = []
    for r in range(2, ws.max_row + 1):
        raw_symbol = ws.cell(r, symbol_col_idx).value
        if raw_symbol is None:
            continue

        raw_symbol = str(raw_symbol).strip()
        if not raw_symbol:
            continue

        binance_symbol = normalize_symbol(raw_symbol)
        base_asset = binance_symbol[:-4] if binance_symbol.endswith("USDT") else raw_symbol.upper()

        rows.append({
            "input_symbol": raw_symbol.upper(),
            "symbol": binance_symbol,
            "base_asset": base_asset,
        })

    return rows


# =========================
# Binance API
# =========================
def get_spot_exchange_info() -> Dict[str, Any]:
    return get_json(f"{SPOT_BASE}/api/v3/exchangeInfo")


def get_futures_exchange_info() -> Dict[str, Any]:
    return get_json(f"{FUTURES_BASE}/fapi/v1/exchangeInfo")


def build_valid_symbol_sets() -> tuple[set[str], set[str]]:
    spot_info = get_spot_exchange_info()
    futures_info = get_futures_exchange_info()

    valid_spot = set()
    valid_futures = set()

    for s in spot_info.get("symbols", []):
        if s.get("status") == "TRADING":
            valid_spot.add(s["symbol"])

    for s in futures_info.get("symbols", []):
        if s.get("status") == "TRADING" and s.get("contractType") == "PERPETUAL":
            valid_futures.add(s["symbol"])

    return valid_spot, valid_futures


def get_spot_24h_all() -> Dict[str, Dict[str, Any]]:
    data = get_json(f"{SPOT_BASE}/api/v3/ticker/24hr")
    return {item["symbol"]: item for item in data}


def get_spot_klines(symbol: str, interval: str = "1d", limit: int = 8) -> List[List[Any]]:
    return get_json(
        f"{SPOT_BASE}/api/v3/klines",
        params={"symbol": symbol, "interval": interval, "limit": limit},
    )


def calc_price_change_7d(symbol: str) -> Optional[float]:
    klines = get_spot_klines(symbol, interval="1d", limit=8)
    if len(klines) < 8:
        return None

    old_close = safe_float(klines[0][4])
    latest_close = safe_float(klines[-1][4])
    return pct_change(latest_close, old_close)


def get_current_open_interest(symbol: str) -> Optional[float]:
    data = get_json(
        f"{FUTURES_BASE}/fapi/v1/openInterest",
        params={"symbol": symbol},
    )
    return safe_float(data.get("openInterest"))


def get_open_interest_hist(symbol: str, period: str = "1d", limit: int = 8) -> List[Dict[str, Any]]:
    return get_json(
        f"{FUTURES_BASE}/futures/data/openInterestHist",
        params={"symbol": symbol, "period": period, "limit": limit},
    )


def calc_oi_changes(symbol: str) -> Dict[str, Optional[float]]:
    hist = get_open_interest_hist(symbol, period="1d", limit=8)
    oi_series = [safe_float(x.get("sumOpenInterest")) for x in hist]
    oi_series = [x for x in oi_series if x is not None]

    result = {
        "oi_change_24h_pct": None,
        "oi_change_7d_pct": None,
    }

    if len(oi_series) >= 2:
        result["oi_change_24h_pct"] = pct_change(oi_series[-1], oi_series[-2])

    if len(oi_series) >= 8:
        result["oi_change_7d_pct"] = pct_change(oi_series[-1], oi_series[0])

    return result


# =========================
# 单个 symbol 处理
# =========================
def analyze_symbol(
    row: Dict[str, str],
    spot_24h_map: Dict[str, Dict[str, Any]],
    valid_spot: set[str],
    valid_futures: set[str],
) -> Dict[str, Any]:
    symbol = row["symbol"]
    input_symbol = row["input_symbol"]
    base_asset = row["base_asset"]

    current_price_num: Optional[float] = None
    current_oi_num: Optional[float] = None
    current_oi_value_num: Optional[float] = None

    result = {
        "输入代币": input_symbol,
        "币安交易对": symbol,
        "代币简称": base_asset,
        "现货已上线": "是" if symbol in valid_spot else "否",
        "合约已上线": "是" if symbol in valid_futures else "否",
        "当前价格": None,
        "24小时涨跌幅": None,
        "7天涨跌幅": None,
        "24小时成交额": None,
        "当前持仓量(OI)": None,     # 代币数量
        "当前持仓金额": None,        # OI数量 × 当前价格
        "OI 24小时涨跌幅": None,
        "OI 7天涨跌幅": None,
        "状态": "正常",
    }

    # 现货数据
    if symbol in valid_spot:
        ticker = spot_24h_map.get(symbol, {})
        current_price_num = safe_float(ticker.get("lastPrice"))

        result["当前价格"] = format_price(current_price_num)
        result["24小时涨跌幅"] = pct_change(
            current_price_num,
            safe_float(ticker.get("openPrice"))
        )
        result["24小时成交额"] = format_large_number(ticker.get("quoteVolume"))

        try:
            result["7天涨跌幅"] = calc_price_change_7d(symbol)
        except Exception as e:
            result["状态"] = f"现货7天数据错误: {e}"
    else:
        result["状态"] = "未上线现货"

    # 合约 OI 数据
    if symbol in valid_futures:
        try:
            current_oi_num = get_current_open_interest(symbol)
            oi_changes = calc_oi_changes(symbol)

            result["当前持仓量(OI)"] = format_large_number(current_oi_num)
            result["OI 24小时涨跌幅"] = oi_changes["oi_change_24h_pct"]
            result["OI 7天涨跌幅"] = oi_changes["oi_change_7d_pct"]

            if current_oi_num is not None and current_price_num is not None:
                current_oi_value_num = current_oi_num * current_price_num
                result["当前持仓金额"] = format_large_number(current_oi_value_num)

        except Exception as e:
            if result["状态"] == "正常":
                result["状态"] = f"合约OI数据错误: {e}"
            else:
                result["状态"] += f" | 合约OI数据错误: {e}"
    else:
        if result["状态"] == "正常":
            result["状态"] = "未上线合约"

    return result


# =========================
# Excel 样式
# =========================
def apply_color_formatting(ws) -> None:
    target_columns = {
        "24小时涨跌幅": "G",
        "7天涨跌幅": "H",
        "OI 24小时涨跌幅": "L",
        "OI 7天涨跌幅": "M",
    }

    green_font = Font(color="008000")
    red_font = Font(color="FF0000")

    max_row = ws.max_row

    for _, col_letter in target_columns.items():
        cell_range = f"{col_letter}2:{col_letter}{max_row}"

        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                font=green_font,
            )
        )

        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                font=red_font,
            )
        )


# =========================
# 导出 Excel
# =========================
def write_results_to_excel(rows: List[Dict[str, Any]], output_file: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "分析结果"

    headers = [
        "输入代币",
        "币安交易对",
        "代币简称",
        "现货已上线",
        "合约已上线",
        "当前价格",
        "24小时涨跌幅",
        "7天涨跌幅",
        "24小时成交额",
        "当前持仓量(OI)",
        "当前持仓金额",
        "OI 24小时涨跌幅",
        "OI 7天涨跌幅",
        "状态",
    ]

    ws.append(headers)

    for row in rows:
        ws.append([row.get(h) for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    percent_columns = {
        "24小时涨跌幅",
        "7天涨跌幅",
        "OI 24小时涨跌幅",
        "OI 7天涨跌幅",
    }

    header_index = {name: idx + 1 for idx, name in enumerate(headers)}

    for col_name in percent_columns:
        col_idx = header_index[col_name]
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"

    center_columns = ["A", "B", "C", "D", "E", "N"]
    for col in center_columns:
        for r in range(2, ws.max_row + 1):
            ws[f"{col}{r}"].alignment = center

    widths = {
        "A": 12,
        "B": 14,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 16,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 16,
        "K": 16,
        "L": 16,
        "M": 16,
        "N": 32,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    apply_color_formatting(ws)
    wb.save(output_file)


# =========================
# 主程序
# =========================
def main():
    print("1) 读取 Excel 中的 symbol ...")
    input_rows = read_symbols_from_excel(INPUT_FILE, INPUT_SHEET_NAME)
    print(f"读取到 {len(input_rows)} 个 symbol")

    print("2) 拉取 Binance 基础信息 ...")
    valid_spot, valid_futures = build_valid_symbol_sets()
    spot_24h_map = get_spot_24h_all()

    print("3) 开始分析 symbol ...")
    results = []
    total = len(input_rows)

    for i, row in enumerate(input_rows, start=1):
        print(f"[{i}/{total}] {row['symbol']}")
        try:
            result = analyze_symbol(row, spot_24h_map, valid_spot, valid_futures)
        except Exception as e:
            result = {
                "输入代币": row["input_symbol"],
                "币安交易对": row["symbol"],
                "代币简称": row["base_asset"],
                "现货已上线": "是" if row["symbol"] in valid_spot else "否",
                "合约已上线": "是" if row["symbol"] in valid_futures else "否",
                "当前价格": None,
                "24小时涨跌幅": None,
                "7天涨跌幅": None,
                "24小时成交额": None,
                "当前持仓量(OI)": None,
                "当前持仓金额": None,
                "OI 24小时涨跌幅": None,
                "OI 7天涨跌幅": None,
                "状态": f"错误: {e}",
            }

        results.append(result)
        time.sleep(0.05)

    print("4) 导出结果 Excel ...")
    write_results_to_excel(results, OUTPUT_FILE)

    print(f"完成，输出文件：{OUTPUT_FILE}")


if __name__ == "__main__":
    main()